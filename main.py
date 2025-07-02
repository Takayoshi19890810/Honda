import pandas as pd
import time
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from datetime import datetime, timedelta
import re
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import json # gspreadの認証情報のため

# ✅ 現在時刻（JST） - 全ニュースソースで使用
now = datetime.utcnow() + timedelta(hours=9)

# ✅ 検索キーワード（複数） - MSNニュースで使用
KEYWORDS_MSN = ["Honda", "HONDA", "ホンダ"]

# ✅ Google/Yahoo!ニュース用の単一キーワード
KEYWORD_SINGLE = "ホンダ"

# ✅ スプレッドシート設定
SPREADSHEET_ID = "1AwwMGKMHfduwPkrtsik40lkO1z1T8IU_yd41ku-yPi8"

def format_datetime(dt_obj):
    """datetimeオブジェクトを指定されたフォーマットの文字列に変換するヘルパー関数"""
    return dt_obj.strftime("%Y/%m/%d %H:%M")

def get_google_news_with_selenium(keyword: str) -> list[dict]:
    """Googleニュースから記事を取得する関数"""
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    url = f"https://news.google.com/search?q={keyword}&hl=ja&gl=JP&ceid=JP:ja"
    driver.get(url)
    time.sleep(5)
    # スクロールしてより多くの記事をロード
    for _ in range(3):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)

    soup = BeautifulSoup(driver.page_source, "html.parser")
    driver.quit()

    articles = soup.find_all("article")
    data = []
    for article in articles:
        try:
            a_tag = article.select_one("a.JtKRv")
            time_tag = article.select_one("time.hvbAAd")
            source_tag = article.select_one("div.vr1PYe")
            title = a_tag.text.strip()
            href = a_tag.get("href")
            # 相対URLを絶対URLに変換
            url = "https://news.google.com" + href[1:] if href.startswith("./") else href
            # 日時をJSTに変換
            dt = datetime.strptime(time_tag.get("datetime"), "%Y-%m-%dT%H:%M:%SZ") + timedelta(hours=9)
            pub_date = format_datetime(dt)
            source = source_tag.text.strip() if source_tag else "N/A"
            data.append({"タイトル": title, "URL": url, "投稿日": pub_date, "引用元": source})
        except Exception as e:
            # print(f"⚠️ Google記事処理エラー: {e}") # デバッグ用
            continue
    print(f"✅ Googleニュース件数: {len(data)} 件")
    return data

def get_yahoo_news_with_selenium(keyword: str) -> list[dict]:
    """Yahoo!ニュースから記事を取得する関数"""
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    search_url = f"https://news.yahoo.co.jp/search?p={keyword}&ei=utf-8"
    driver.get(search_url)
    time.sleep(5)

    soup = BeautifulSoup(driver.page_source, "html.parser")
    driver.quit()
    # 記事コンテナのセレクターを調整
    articles = soup.find_all("li", class_=re.compile("sc-1u4589e-0"))
    data = []

    for article in articles:
        try:
            title_tag = article.find("div", class_=re.compile("sc-3ls169-0"))
            title = title_tag.text.strip() if title_tag else ""
            link_tag = article.find("a", href=True)
            url = link_tag["href"] if link_tag else ""
            time_tag = article.find("time")
            date_str = time_tag.text.strip() if time_tag else ""
            formatted_date = ""
            if date_str:
                date_str = re.sub(r'\([月火水木金土日]\)', '', date_str).strip() # 曜日を削除
                try:
                    dt_obj = datetime.strptime(date_str, "%Y/%m/%d %H:%M")
                    formatted_date = format_datetime(dt_obj)
                except ValueError:
                    formatted_date = date_str # フォーマットできない場合はそのまま
            source_text = "Yahoo!"
            if title and url:
                data.append({
                    "タイトル": title,
                    "URL": url,
                    "投稿日": formatted_date if formatted_date else "取得不可",
                    "引用元": source_text
                })
        except Exception as e:
            # print(f"⚠️ Yahoo!記事処理エラー: {e}") # デバッグ用
            continue

    print(f"✅ Yahoo!ニュース件数: {len(data)} 件")
    return data

def get_msn_news_with_selenium(keywords: list[str]) -> list[dict]:
    """MSNニュースから記事を取得する関数（複数のキーワードに対応）"""
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")

    # driverの初期化はループの外で行い、キーワードごとにページをロード
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    all_msn_data = []

    for keyword in keywords:
        print(f"🔍 MSNニュース - 処理中: {keyword}")
        search_url = f'https://www.bing.com/news/search?q={keyword}&qft=sortbydate%3d"1"&form=YFNR'
        driver.get(search_url)
        time.sleep(5) # ページが完全にロードされるように少し長めに待つ

        soup = BeautifulSoup(driver.page_source, "html.parser")

        # div.news-card を探し、データ属性から情報を抽出
        for card in soup.select('div.news-card'):
            title = card.get("data-title", "").strip()
            url = card.get("data-url", "").strip()
            source = card.get("data-author", "").strip()

            pub_time_obj = None
            pub_label = ""

            # 投稿時間をaria-label属性から取得
            pub_tag = card.find("span", attrs={"aria-label": True})
            if pub_tag and pub_tag.has_attr("aria-label"):
                pub_label = pub_tag["aria-label"].strip()

            # 相対時間表記をdatetimeオブジェクトに変換するロジック
            if "分前" in pub_label:
                m = re.search(r"(\d+)", pub_label)
                if m:
                    pub_time_obj = now - timedelta(minutes=int(m.group(1)))
            elif "時間前" in pub_label:
                h = re.search(r"(\d+)", pub_label)
                if h:
                    pub_time_obj = now - timedelta(hours=int(h.group(1)))
            elif "日前" in pub_label:
                d = re.search(r"(\d+)", pub_label)
                if d:
                    pub_time_obj = now - timedelta(days=int(d.group(1)))
            elif re.match(r'\d+月\d+日', pub_label): # 例: 1月1日
                try:
                    # 年は現在の年を使用すると仮定
                    pub_time_obj = datetime.strptime(f"{now.year}年{pub_label}", "%Y年%m月%d日")
                except:
                    pub_time_obj = None
            elif re.match(r'\d{4}/\d{1,2}/\d{1,2}', pub_label): # 例: 2024/01/01
                try:
                    pub_time_obj = datetime.strptime(pub_label, "%Y/%m/%d")
                except:
                    pub_time_obj = None
            elif re.match(r'\d{1,2}:\d{2}', pub_label): # 例: 15:30 (今日の日付と結合)
                try:
                    t = datetime.strptime(pub_label, "%H:%M").time()
                    pub_time_obj = datetime.combine(now.date(), t)
                except:
                    pub_time_obj = None

            pub_date = pub_time_obj.strftime("%Y/%m/%d %H:%M") if pub_time_obj else pub_label

            if title and url:
                all_msn_data.append({
                    "タイトル": title,
                    "URL": url,
                    "投稿日": pub_date,
                    "引用元": source if source else "MSN"
                })
        # print(f"✅ MSNニュース - キーワード '{keyword}' で {len(all_msn_data)} 件の記事を発見。") # デバッグ用
    
    driver.quit() # 全キーワード処理後に一度だけドライバーを終了
    print(f"✅ MSNニュース総件数: {len(all_msn_data)} 件")
    return all_msn_data

def write_to_spreadsheet(articles: list[dict], spreadsheet_id: str, worksheet_name: str):
    """記事データをGoogleスプレッドシートに書き込む関数"""
    # 環境変数から認証情報を取得
    credentials_json_str = os.environ.get('GCP_SERVICE_ACCOUNT_KEY')
    # 環境変数がない場合はファイルから読み込む（ローカル開発用）
    if credentials_json_str:
        credentials = json.loads(credentials_json_str)
    else:
        # ローカルでのテスト時に credentials.json が必要
        # 本番環境では環境変数 GCP_SERVICE_ACCOUNT_KEY にJSON文字列を設定してください
        try:
            with open('credentials.json', 'r') as f:
                credentials = json.load(f)
        except FileNotFoundError:
            raise RuntimeError("認証ファイル 'credentials.json' が見つからないか、GCP_SERVICE_ACCOUNT_KEY 環境変数が設定されていません。")

    gc = gspread.service_account_from_dict(credentials)

    for attempt in range(5): # 複数回リトライ
        try:
            sh = gc.open_by_key(spreadsheet_id)
            try:
                worksheet = sh.worksheet(worksheet_name)
            except gspread.exceptions.WorksheetNotFound:
                # ワークシートが存在しない場合は新規作成しヘッダーを追加
                worksheet = sh.add_worksheet(title=worksheet_name, rows="1", cols="4")
                worksheet.append_row(['タイトル', 'URL', '投稿日', '引用元'])

            existing_data = worksheet.get_all_values()
            # 既存のURLをセットに格納して重複チェックを高速化
            existing_urls = set(row[1] for row in existing_data[1:] if len(row) > 1)

            # 新しいデータのみをフィルタリング
            new_data = [[a['タイトル'], a['URL'], a['投稿日'], a['引用元']] for a in articles if a['URL'] not in existing_urls]
            
            if new_data:
                # 新しいデータをスプレッドシートに追記
                worksheet.append_rows(new_data, value_input_option='USER_ENTERED')
                print(f"✅ {len(new_data)}件をスプレッドシートに追記しました。")
            else:
                print("⚠️ 追記すべき新しいデータはありません。")
            return # 成功したら関数を終了
        except gspread.exceptions.APIError as e:
            print(f"⚠️ Google API Error (attempt {attempt + 1}/5): {e}")
            time.sleep(5 + random.random() * 5) # リトライ前に待機
    
    # 5回試行しても成功しない場合はエラー
    raise RuntimeError("❌ Googleスプレッドシートへの書き込みに失敗しました（5回試行しても成功せず）")

if __name__ == "__main__":
    print("\n--- Google News ---")
    google_news_articles = get_google_news_with_selenium(KEYWORD_SINGLE)
    if google_news_articles:
        write_to_spreadsheet(google_news_articles, SPREADSHEET_ID, "Google")
    
    print("\n--- Yahoo! News ---")
    yahoo_news_articles = get_yahoo_news_with_selenium(KEYWORD_SINGLE)
    if yahoo_news_articles:
        write_to_spreadsheet(yahoo_news_articles, SPREADSHEET_ID, "Yahoo")

    # MSNニュースの処理は、キーワードリストで渡す
    print("\n--- MSN News ---")
    msn_news_articles = get_msn_news_with_selenium(KEYWORDS_MSN)
    # ここで重複排除 (get_msn_news_with_selenium 内ではキーワードごとの重複排除はしていないため)
    df_msn = pd.DataFrame(msn_news_articles)
    if not df_msn.empty:
        df_msn.drop_duplicates(subset=["URL"], inplace=True)
        write_to_spreadsheet(df_msn.to_dict('records'), SPREADSHEET_ID, "MSN")
    else:
        print("⚠️ MSNニュースの追記すべき新しいデータはありません。")
    
    print("\n--- 全てのニュースソースの抽出とスプレッドシートへの書き込みが完了しました ---")

    # ローカルExcelファイルへの書き出し処理（元のMSN専用スクリプトから統合）
    # 全てのデータを統合して重複排除
    combined_all_data = google_news_articles + yahoo_news_articles + msn_news_articles # 型がdictのリストであることを前提

    # combined_all_data を DataFrame に変換して重複排除
    df_combined = pd.DataFrame(combined_all_data)
    if not df_combined.empty:
        df_combined.drop_duplicates(subset=["URL"], inplace=True)
    
        output_file = "all_news_summary.xlsx" # 統合された出力ファイル名
        df_combined.to_excel(output_file, index=False)

        # フィルタ付きテーブル追加
        try:
            wb = load_workbook(output_file)
            ws = wb.active

            end_row = ws.max_row
            end_col = ws.max_column
            end_col_letter = chr(ord('A') + end_col - 1)
            table_range = f"A1:{end_col_letter}{end_row}"

            table = Table(displayName="CombinedNewsTable", ref=table_range)
            style = TableStyleInfo(name="TableStyleLight1", showRowStripes=False)
            table.tableStyleInfo = style
            ws.add_table(table)
            wb.save(output_file)
            print("✅ 統合Excelファイルにフィルタ付きテーブルを追加しました。")
        except Exception as e:
            print(f"⚠️ 統合Excelファイルのフィルタ設定エラー: {e}")

        print(f"✅ 全ニュースソースの抽出完了: {output_file}（件数: {len(df_combined)}）")
    else:
        print("⚠️ 全てのニュースソースで新しいデータがありませんでした。")
